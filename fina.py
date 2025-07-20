import os
import json
import time
import requests
from openpyxl import load_workbook

# === CONFIG === (confidential information is removed or masked)
LOGIN = "your_username"
PASSWORD = "your_password"
BASE_URL = "http://your.api.endpoint"
TOKEN_FILE = "token2.json"
TOKEN_VALID_FOR_SECONDS = 36 * 3600
SELF_COST_DATE = "2024-01-01"
EXCEL_FILE = "bigExcel.xlsx"

PRICE_IDS = {
    "company": 8,
    "discounted": 9,
    "reference": 5
}

# === AUTH ===
def authenticate():
    url = f"{BASE_URL}/api/authentication/authenticate"
    payload = {"login": LOGIN, "password": PASSWORD}
    print("🔑 Getting token...")
    response = requests.post(url, json=payload)
    response.raise_for_status()
    data = response.json()

    if "token" not in data or not data["token"]:
        raise Exception("❌ No token received")

    with open(TOKEN_FILE, "w") as f:
        json.dump({"token": data["token"], "timestamp": int(time.time())}, f)
    return data["token"]

def get_token():
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "r") as f:
            info = json.load(f)
            if int(time.time()) - info["timestamp"] < TOKEN_VALID_FOR_SECONDS:
                return info["token"]
    return authenticate()

# === FETCH PRIMARY PRODUCT DATA ===
def get_products_complete():
    token = get_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    products = requests.get(f"{BASE_URL}/api/operation/getProducts", headers=headers).json().get("products", [])
    ids = [p["id"] for p in products]

    prices = requests.get(f"{BASE_URL}/api/operation/getProductPrices", headers=headers).json().get("prices", [])
    price_map = {p["product_id"]: {
        "price": p["price"],
        "discount_price": p["discount_price"]
    } for p in prices}

    cost_resp = requests.post(
        f"{BASE_URL}/api/operation/getProductsSelfCost",
        headers=headers,
        json={"prods": ids, "date": SELF_COST_DATE}
    ).json()
    cost_map = {c["id"]: c["cost"] for c in cost_resp.get("cost_info", [])}

    product_map = {}
    for p in products:
        code = str(p["code"]).strip()
        pid = p["id"]
        product_map[code] = {
            "price": price_map.get(pid, {}).get("price"),
            "discount_price": price_map.get(pid, {}).get("discount_price"),
            "self_cost": cost_map.get(pid)
        }

    return product_map, products

# === FETCH ADVANCED PRICE DATA ===
def get_prices_by_type(token, product_ids, price_id):
    url = f"{BASE_URL}/api/operation/getProductPricesAdvance"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    prices = {}
    chunks = [product_ids[i:i + 100] for i in range(0, len(product_ids), 100)]

    for chunk in chunks:
        response = requests.post(url, headers=headers, json={"prods": chunk, "price": price_id})
        response.raise_for_status()
        for p in response.json().get("prices", []):
            prices[p["product_id"]] = p.get("price")

    return prices

def fetch_all_prices(token, product_ids):
    return {
        "company": get_prices_by_type(token, product_ids, PRICE_IDS["company"]),
        "discounted": get_prices_by_type(token, product_ids, PRICE_IDS["discounted"]),
        "reference": get_prices_by_type(token, product_ids, PRICE_IDS["reference"])
    }

# === EXCEL UPDATE ===
def update_excel(product_map, prices, code_to_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    total = matched = 0

    print("\n📊 Updating Excel...")

    for row in range(2, ws.max_row + 1):
        val = ws[f"A{row}"].value
        if not val:
            continue

        key = str(val).strip()
        if key.upper() == "STOP":
            print(f"🛑 Found STOP at row {row}. Ending.")
            break

        total += 1
        product = product_map.get(key)
        pid = code_to_id.get(key)

        if product:
            ws[f"Q{row}"] = product.get("price")
            ws[f"R{row}"] = product.get("discount_price")
            ws[f"S{row}"] = product.get("self_cost")
        else:
            ws[f"Q{row}"] = -1
            ws[f"R{row}"] = -1
            ws[f"S{row}"] = -1

        if pid:
            ws[f"T{row}"] = prices["company"].get(pid, -1)
            ws[f"U{row}"] = prices["discounted"].get(pid, -1)
            ws[f"V{row}"] = prices["reference"].get(pid, -1)
            matched += 1
        else:
            ws[f"T{row}"] = -1
            ws[f"U{row}"] = -1
            ws[f"V{row}"] = -1
            print(f"❌ Row {row}: '{key}' not found")

    wb.save(EXCEL_FILE)
    print(f"\n✅ Done. Checked: {total} | Matched: {matched}")

# === MAIN ===
def main():
    token = get_token()
    product_map, products = get_products_complete()
    product_ids = [p["id"] for p in products]
    code_to_id = {str(p["code"]).strip(): p["id"] for p in products}
    prices = fetch_all_prices(token, product_ids)
    update_excel(product_map, prices, code_to_id)

if __name__ == "__main__":
    main()
